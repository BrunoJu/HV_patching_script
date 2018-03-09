#ScriptAim: Get data from Hypervisor Excel automatically
#Maintenance: BrunoJu
#Version: 1.0

###################Warning##############################Warning#####################################warning###########################
#Before using this script, please make sure you have sorted the servers by source_server in ervery sheet no matter A-Z or Z-A.       #
#It means to eliminate the in-between blank rows, otherwise the script may not waork as you expected.                                #
######################################################################################################################################

import os,re
from openpyxl import load_workbook


#Insert the Work Excel you do.
filename = raw_input("Input Workbook with Absolute URL: ")
wb = load_workbook(filename = filename)
#wb = load_workbook(filename = '//DWDF219/CPS_BIT_Ops/L2/tracking/Hypervisor Patching/progress_list_mar_12.xlsx')

#Initialize some Vars.
valid_sheets=[]
#source_server list
s_blade_list=[]
#target_server list
t_blade_list=[]
#First colum with server name.
#You must change the number if it changed in Excel.
start_column= 8
#Find the max column of all sheets
max_column= 0


#Filter valid sheets list
for i in wb.sheetnames:
    if i.lstrip().rstrip().lower().find('summary') + i.lstrip().rstrip().lower().find('pool')== -2:
        valid_sheets.append(i)

#Get Hardware_pool source_server and target_server
for sheets_name in valid_sheets:
        specified_sheet=wb[sheets_name]

#Get the max end_column number
        for i in range(start_column,100):
            if specified_sheet['A'+str(i)].value != None:
                ++i
            else:
                if i > max_column:
                    max_column = i


#                    specified_sheet.auto_filter.ref = "A8:O"+str(max_column)
#                    specified_sheet.auto_filter.add_sort_condition("A8:A"+str(max_column))
#                    wb.save('C:\Users\c5258719\Desktop\progress_list_mar_10.xlsx')

#Trun out the vars to create ticket.
        for i in range(start_column,max_column):
            if specified_sheet['D'+str(i)].value == None and \
                specified_sheet['C'+str(i)].value != None:
                s_blade_list.append(specified_sheet['A'+str(i)].value)
                t_blade_list.append(specified_sheet['C'+str(i)].value)



#Get data from Excel successfully now.
#Please check the list if is None when using JS to create ticket.
