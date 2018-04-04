#coding=utf-8
import os,re,time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

#Insert the Work Excel you do.

###################Warning##############################Warning#####################################warning###########################
#Before using this script, please make sure you have sorted the servers by source_server in ervery sheet no matter A-Z or Z-A.       #
#It means to eliminate the in-between blank rows, otherwise the script may not waork as you expected.                                #
######################################################################################################################################

filename = input("Input Workbook with Absolute URL: ")
wb = load_workbook(filename = filename)
#wb = load_workbook(filename = '//DWDF219/CPS_BIT_Ops/L2/tracking/Hypervisor Patching/progress_list_mar_12.xlsx')

#Initialize some Vars.
valid_sheets=[]
#source_server list
s_blade_list=[]
#target_server list
t_blade_list=[]
#First colum with server name.
#You must change the number if it changed in Excel.
start_column= 8
#Find the max column of all sheets
max_column= 0

#Filter valid sheets list
for i in wb.sheetnames:
    if i.lstrip().rstrip().lower().find('summary') + i.lstrip().rstrip().lower().find('pool')== -2:
        valid_sheets.append(i)

#Get Hardware_pool source_server and target_server
for sheets_name in valid_sheets:
        specified_sheet=wb[sheets_name]

#Get the max end_column number
        for i in range(start_column,100):
            if specified_sheet['A'+str(i)].value != None:
                ++i
            else:
                if i > max_column:
                    max_column = i

#                    specified_sheet.auto_filter.ref = "A8:O"+str(max_column)
#                    specified_sheet.auto_filter.add_sort_condition("A8:A"+str(max_column))
#                    wb.save('C:\Users\c5258719\Desktop\progress_list_mar_10.xlsx')

#Trun out the vars to create ticket.
        for i in range(start_column,max_column):
            if specified_sheet['D'+str(i)].value == None and \
                specified_sheet['C'+str(i)].value != None:
                s_blade_list.append(specified_sheet['A'+str(i)].value)
                t_blade_list.append(specified_sheet['C'+str(i)].value)

#Get data from Excel successfully now.
#Please check the list if is None when using JS to create ticket.

print(s_blade_list)
print(t_blade_list)

#use localprofile to open firefox
gmp_index_boolean = False
gmp_ticket_boolean = False
f=open('profile_directory.txt','r')
s=f.readlines()
#profile_directory = r'C:\Users\c5258641\AppData\Roaming\Mozilla\Firefox\Profiles\7870jpat.default'
profile_directory=s[1]
profile = webdriver.FirefoxProfile(profile_directory)
driver = webdriver.Firefox(profile)

while(gmp_index_boolean == False | gmp_ticket_boolean == False):
	driver.get("https://gmp.wdf.sap.corp/")
	index_boolean = WebDriverWait(driver,15).until(EC.title_is(u"GMP - Portal"))
	gmp_index_boolean = index_boolean
	time.sleep(5)
	driver.get("https://gmp.wdf.sap.corp/cgi-bin/off_dispatch.pl/plan/user")
	ticket_boolean = WebDriverWait(driver,10).until(EC.title_is(u"GMP - Personal Worklist"))
	gmp_ticket_boolean = ticket_boolean

print("Page Finishes Loading")


#javascript function
def runJavaScript(s_blade,t_blade):
	#the javascript
	createTicket_js = r'''var s_blade = "'''+s_blade+'''";var t_blade = "'''+t_blade+'''";
    s_blade = s_blade.replace(/\s/ig,'');
    t_blade = t_blade.replace(/\s/ig,'');
	var ticket = document.getElementsByClassName("btn btn-primary fancybox");
    var createdEvent = document.createEvent("MouseEvents");
    createdEvent.initMouseEvent("click", true, true);
    ticket[0].dispatchEvent(createdEvent)
    setTimeout(function(){
    var internalWindow = document.getElementsByTagName("iframe")[0].contentWindow
    var internalDocument = internalWindow.document
    var edit = internalDocument.getElementsByClassName('editcontrol')
    edit[3].children[0].innerText = "Hypervisor patching\\n\\nServer to patch:"+s_blade+"\\nDestination Server:"+t_blade+"\\n\\nPlease follow the documentation:\\nhttps://gmp-twiki.wdf.sap.corp/cgi-bin/twiki/bin/view/A1s/A1S_new/ServiceRequest/CrossProductsServices/HypervisorPatching"
    edit[3].children[0].value = "Hypervisor patching\\n\\nServer to patch:"+s_blade+"\\nDestination Server:"+t_blade+"\\n\\nPlease follow the documentation:\\nhttps://gmp-twiki.wdf.sap.corp/cgi-bin/twiki/bin/view/A1s/A1S_new/ServiceRequest/CrossProductsServices/HypervisorPatching"
    edit[1].children[0].value="M"
    var internalDocument = document.getElementsByTagName("iframe")[0].contentWindow.document
    var btn_search = internalDocument.getElementsByClassName("ui-button-text")
    btn_search[3].dispatchEvent(createdEvent)
    btn_search[5].dispatchEvent(createdEvent)
    btn_search[3].dispatchEvent(createdEvent)
    btn_search[9].dispatchEvent(createdEvent)
    btn_search[5].dispatchEvent(createdEvent)
    var inventory = internalDocument.getElementsByClassName('glyphicon glyphicon-equalizer')
    inventory[1].dispatchEvent(createdEvent)
    setTimeout(function(){
    internalWindow.$('#inv_search_input').val(s_blade)
    internalWindow.inventory_search()
    setTimeout(function(){
    var table = internalDocument.getElementsByClassName("display table table-striped table-bordered table-hover table-condensed dataTable")
    if(s_blade.indexOf("rsa")==0){
    var input_SName = table[1].getElementsByTagName("input")[1]
    }else{
    var input_SName = table[0].getElementsByTagName("input")[1]
    }
    input_SName.checked=true
    internalWindow.inventory_select()
    setTimeout(function(){
    internalWindow.suggest_field_for_inventory("area")
    internalWindow.suggest_field_for_inventory("landscape")
    internalWindow.suggest_field_for_inventory("usage_area")
    setTimeout(function(){
    internalWindow.save_base()
    },7000)
    },3000)
    },15000)
    },8000)
    },8000)
	'''
	driver.execute_script(createTicket_js)

#main function
if len(s_blade_list) != 0:
	for (s_blade, t_blade) in zip(s_blade_list,t_blade_list):
		runJavaScript(s_blade,t_blade)
		time.sleep(60)
else:
	print("Can't find source server, please retry !")

driver.quit()
